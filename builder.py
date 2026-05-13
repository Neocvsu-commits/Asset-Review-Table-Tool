"""扫描资产目录、调用渲染器、写出 xlsx 表格。"""

from __future__ import annotations

import random
from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import Callable, Iterable

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from renderer import render_one
from utils import (
    find_model_files,
    find_render_model,
    human_size,
    parse_basic_csv,
    skeleton_animation,
    sum_loose_texture_bytes,
    texture_size_summary,
)


HEADERS: tuple[str, ...] = (
    "资产名称", "中文名称", "骨骼", "动画", "动画类型",
    "模型截图", "三角面数", "材质球数量", "贴图数量",
    "贴图尺寸(去重摘要)", "贴图文件总大小(散文件)",
    "FBX文件大小", "GLB文件大小", "是否入库",
)
IMG_COL = 6
THUMB_MAX_W = 220
ROW_HEIGHT = 120


@dataclass
class _Row:
    name: str
    cn: str
    skeleton: str
    animation: str
    anim_type: str
    tris: str
    mats: str
    tex_n: str
    tex_sum: str
    tex_total: str
    fbx_sz: str
    glb_sz: str
    png: Path | None


def _gather_candidates(roots: Iterable[Path], log: Callable[[str], None]) -> list[Path]:
    items: list[Path] = []
    for root in roots:
        if not root.is_dir():
            log(f"警告: 目录不存在，已跳过: {root}")
            continue
        for child in sorted(root.iterdir()):
            if not child.is_dir():
                continue
            if not any(child.glob("*_BasicInformation.csv")):
                continue
            if find_render_model(child) is None:
                continue
            items.append(child)
    return items


def _deduplicate(folders: list[Path], log: Callable[[str], None]) -> list[Path]:
    seen: set[str] = set()
    kept: list[Path] = []
    for folder in folders:
        csv_path = next(folder.glob("*_BasicInformation.csv"))
        name = parse_basic_csv(csv_path).get("资产名称", folder.name)
        if name in seen:
            log(f"跳过重复资产名: {name} <- {folder}")
            continue
        seen.add(name)
        kept.append(folder)
    return kept


def _row_from_folder(folder: Path, thumb_dir: Path, blender: Path, hdr: Path | None, error_log: Path) -> _Row:
    csv_path = next(folder.glob("*_BasicInformation.csv"))
    info = parse_basic_csv(csv_path)
    name = info.get("资产名称", folder.name)
    fbx, glb = find_model_files(folder)
    skel, anim, anim_type = skeleton_animation(info)

    png_out = thumb_dir / f"{name}.png"
    ok = False
    model = find_render_model(folder)
    if model:
        ok, log_text = render_one(model, png_out, blender, hdr=hdr)
        if not ok and log_text:
            try:
                with error_log.open("a", encoding="utf-8") as fp:
                    fp.write(f"\n==== {model} -> {png_out} ====\n{log_text}\n")
            except OSError:
                pass

    return _Row(
        name=name,
        cn=info.get("资产中文名称", ""),
        skeleton=skel,
        animation=anim,
        anim_type=anim_type,
        tris=info.get("三角面数", ""),
        mats=info.get("材质球数量", ""),
        tex_n=info.get("贴图数量", ""),
        tex_sum=texture_size_summary(csv_path),
        tex_total=human_size(sum_loose_texture_bytes(folder)),
        fbx_sz=human_size(fbx.stat().st_size) if fbx else "",
        glb_sz=human_size(glb.stat().st_size) if glb else "",
        png=png_out if ok else None,
    )


def _write_workbook(rows: list[_Row], xlsx_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "校验样例"
    for col, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.column_dimensions[get_column_letter(IMG_COL)].width = 28
    ws.column_dimensions[get_column_letter(len(HEADERS))].width = 12

    for i, r in enumerate(rows, start=2):
        ws.row_dimensions[i].height = ROW_HEIGHT
        values = [
            r.name, r.cn, r.skeleton, r.animation, r.anim_type,
            None,  # 截图列稍后用 add_image 填入
            r.tris, r.mats, r.tex_n, r.tex_sum, r.tex_total,
            r.fbx_sz, r.glb_sz, "",
        ]
        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=i, column=col, value=val)
            if col != IMG_COL:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        if r.png and r.png.exists():
            img = XLImage(str(r.png))
            if img.width > THUMB_MAX_W > 0:
                scale = THUMB_MAX_W / img.width
                img.height = int(img.height * scale)
                img.width = THUMB_MAX_W
            ws.add_image(img, f"{get_column_letter(IMG_COL)}{i}")
        else:
            ws.cell(row=i, column=IMG_COL, value="(渲染失败或无可用模型)")

    wb.save(xlsx_path)


def _xlsx_filename(out_dir: Path, root_count: int, limit: int | None) -> Path:
    today = date.today().strftime("%Y%m%d")
    if limit:
        return out_dir / f"资产review流程测试样例_{today}.xlsx"
    if root_count > 1:
        return out_dir / f"资产review归档_合并_{today}.xlsx"
    return out_dir / f"资产review归档_全量_{today}.xlsx"


def build_report(
    roots: list[Path],
    out_dir: Path,
    blender: Path,
    *,
    hdr: Path | None = None,
    limit: int | None = None,
    seed: int = 20260512,
    log: Callable[[str], None] = print,
) -> Path:
    """执行完整流程：扫描 → 去重 → 渲染 → 写表。返回生成的 xlsx 路径。"""
    roots = [r.expanduser().resolve() for r in roots]
    out_dir = out_dir.expanduser().resolve()
    out_dir.mkdir(parents=True, exist_ok=True)
    thumb_dir = out_dir / "thumbnails"
    thumb_dir.mkdir(parents=True, exist_ok=True)

    error_log = out_dir / "_render_errors.log"
    if error_log.exists():
        try:
            error_log.unlink()
        except OSError:
            pass

    raw = _gather_candidates(roots, log)
    if not raw:
        raise RuntimeError("未找到任何包含 *_BasicInformation.csv 与 .glb/.fbx 的资产文件夹")

    candidates = _deduplicate(raw, log)
    if not candidates:
        raise RuntimeError("去重后无可用资产")

    if limit and limit > 0:
        if len(candidates) < limit:
            raise RuntimeError(f"需要 {limit} 个资产，仅找到 {len(candidates)} 个")
        picked = random.Random(seed).sample(candidates, limit)
    else:
        picked = sorted(candidates, key=lambda p: p.name.lower())

    log(f"准备处理 {len(picked)} 个资产...")
    rows: list[_Row] = []
    for i, folder in enumerate(sorted(picked, key=lambda p: p.name), start=1):
        log(f"[{i}/{len(picked)}] {folder.name}")
        rows.append(_row_from_folder(folder, thumb_dir, blender, hdr, error_log))

    xlsx_path = _xlsx_filename(out_dir, len(roots), limit)
    _write_workbook(rows, xlsx_path)
    log(f"已生成: {xlsx_path}（{len(rows)} 行）")
    return xlsx_path
